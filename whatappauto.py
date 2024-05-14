import pywhatkit
import random
import openpyxl
import time

wrkbk = openpyxl.load_workbook("contact2send.xlsx")
sh = wrkbk.active

for i in range(1, sh.max_row+1):
#cell_obj = sh.cell(row=i, column=j)
	message= """Hi {(sh.cell(row=i, column=1)).value},
	Your candudate id{(sh.cell(row=i, column=3)).value}and selected to python course """
	
	number = (sh.cell(row=i, column=2)).value

#print(cell_obj.value)
	'''pywhatkit.sendwhatmsg_instantly('+', message)
	time.sleep(random.randrange(0, 3))'''
		


'''
for i in range(0,len(a)):
	pywhatkit.sendwhatmsg_instantly(a[i],c[0])
	sleep(random.randrange(0, 3))'''

